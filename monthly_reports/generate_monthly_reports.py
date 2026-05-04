"""
CAPS Monthly Performance Report Generator — Simple Edition
==========================================================
A plain, executive-friendly one-pager.

Headline KPIs:    Submitted · Interview · BAFO · Awards (by ITA date)
Plus dollar values: Bid Value submitted · Revenue generated (by ITA date)

Charts:
  · Service category breakdown (count + $ value)
  · Top states for the month
  · 6-month trend: submissions count + revenue ($) from ITA-dated deals

Single primary color (deep navy) + a single soft accent (teal).
No funnel, no gauge, no chartjunk.

Usage:
    python generate_monthly_reports.py
    python generate_monthly_reports.py --month 2026-04
"""

import os, sys, io, argparse, json, re
from datetime import datetime, date
from dateutil.relativedelta import relativedelta
from collections import defaultdict, OrderedDict

import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

from svglib.svglib import svg2rlg
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    Image, HRFlowable
)

# ─────────────────────────────────────────────
#  Color palette  — strict: navy + one accent
# ─────────────────────────────────────────────
NAVY    = '#0F1B2E'   # primary headlines & KPI numbers
INK     = '#1E293B'   # body text
MUTED   = '#64748B'   # labels, sub-text
FAINT   = '#94A3B8'   # footnotes, footer
RULE    = '#E2E8F0'   # rules
STRIPE  = '#F8FAFC'   # zebra stripe
WHITE   = '#FFFFFF'
ACCENT  = '#0EA5A4'   # teal — bars, accents (used sparingly)

def rl(h): return colors.HexColor(h)

# ─────────────────────────────────────────────
#  Paths & teams
# ─────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
PROJECT_DIR = os.path.dirname(os.path.dirname(SCRIPT_DIR))
EXCEL_FILE  = os.path.join(PROJECT_DIR, 'CAPS_RFP_Dashboard_Dataset.xlsx')
LOGO_PATH   = os.path.join(PROJECT_DIR, 'image.png')

TEAMS = {
    'Team Alpha'  : ['Ashutosh Chauhan', 'Madhav Mitruka'],
    'Team Kairoz' : ['Jyothsna Nanda kishore'],
    'Team D'      : ['Savya Atrey', 'Abhishek Zalani', 'Srikrishnadeva D'],
}

def owner_to_team(owner):
    if not owner: return None
    lo = owner.strip().lower()
    for t, owners in TEAMS.items():
        if any(lo == o.lower() for o in owners): return t
    return None

# ─────────────────────────────────────────────
#  Utilities
# ─────────────────────────────────────────────
def parse_date(v):
    if not v: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date):     return v
    try: return datetime.strptime(str(v)[:10], '%Y-%m-%d').date()
    except Exception: return None

def parse_num(v):
    try: return float(str(v or '').replace(',', '').replace('$', '')) if v else 0.0
    except Exception: return 0.0

def money(v):
    if not v: return '—'
    if v >= 1_000_000: return f'${v/1_000_000:.1f}M'
    if v >= 1_000:     return f'${v/1_000:.0f}K'
    return f'${v:,.0f}'


def xml_escape(s):
    """Escape XML special chars for ReportLab Paragraph markup."""
    if s is None: return ''
    return (str(s).replace('&', '&amp;')
                  .replace('<', '&lt;')
                  .replace('>', '&gt;'))

def in_month(d, yr, mo):
    return d is not None and d.year == yr and d.month == mo


def in_period(d, year, months):
    """True if `d` falls in any of the given months in `year`."""
    return d is not None and d.year == year and d.month in months


def quarter_months(q):
    """Q1=[1,2,3], Q2=[4,5,6], Q3=[7,8,9], Q4=[10,11,12]."""
    return list(range((q - 1) * 3 + 1, (q - 1) * 3 + 4))

# ─────────────────────────────────────────────
#  Data loading
# ─────────────────────────────────────────────
def load_data():
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    ws = wb['RFP Data']
    h  = [c.value for c in ws[1]]
    gi = lambda name: h.index(name) if name in h else None
    ix = {k: gi(v) for k, v in {
        'id':'HubSpot ID','rfp_no':'RFP Number',
        'name':'Deal Name','agency':'Agency','state':'Agency State',
        'stage':'Stage','sub':'Submission Date','amount':'Amount ($)',
        'cat':'Service Category','owner':'Owner',
        'iv_date':'Interview Date','bafo':'BAFO Date',
        'ita':'Intent to Award Date','awarded':'Awarded Date',
        'link':'HubSpot Link',
    }.items()}
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        g = lambda k: row[ix[k]] if ix.get(k) is not None else None
        rows.append({
            'id'     : str(g('id')      or ''),
            'rfp_no' : str(g('rfp_no')  or ''),
            'name'   : str(g('name')    or ''),
            'agency' : str(g('agency')  or ''),
            'state'  : str(g('state')   or ''),
            'stage'  : str(g('stage')   or ''),
            'sub'    : parse_date(g('sub')),
            'amount' : parse_num(g('amount')),
            'cat'    : str(g('cat')     or ''),
            'owner'  : str(g('owner')   or ''),
            'iv_date': parse_date(g('iv_date')),
            'bafo'   : parse_date(g('bafo')),
            'ita'    : parse_date(g('ita')),
            'awarded': parse_date(g('awarded')),
            'link'   : str(g('link')    or ''),
            'team'   : owner_to_team(g('owner')),
        })
    return rows

def by_category(deals):
    cats = defaultdict(lambda: {'count': 0, 'amount': 0.0})
    for d in deals:
        for c in [x.strip() for x in (d['cat'] or '').split(';') if x.strip()]:
            cats[c]['count']  += 1
            cats[c]['amount'] += d['amount']
    return OrderedDict(sorted(cats.items(), key=lambda x: -x[1]['count']))

def by_state(deals, top=8):
    sc = defaultdict(int)
    for d in deals:
        s = (d['state'] or '').strip()
        if s:
            sc[s] += 1
    return OrderedDict(sorted(sc.items(), key=lambda x: -x[1])[:top])

def six_months(team_deals, yr, mo):
    """Last 6 months relative to (yr, mo): subs count, sub_value, revenue (ITA date)."""
    out = []
    for off in range(5, -1, -1):
        t = date(yr, mo, 1) - relativedelta(months=off)
        subs = [d for d in team_deals if in_month(d['sub'], t.year, t.month)]
        ita  = [d for d in team_deals if in_month(d['ita'], t.year, t.month)]
        out.append({
            'label'    : t.strftime('%b'),
            'year'     : t.year, 'month': t.month,
            'subs'     : len(subs),
            'sub_value': sum(d['amount'] for d in subs),
            'awards'   : len(ita),
            'revenue'  : sum(d['amount'] for d in ita),
        })
    return out


def four_quarters(team_deals, yr, q):
    """Last 4 quarters relative to (yr, q): subs count, revenue (ITA date)."""
    out = []
    for off in range(3, -1, -1):
        # Walk back `off` quarters from (yr, q)
        target_q_index = (yr * 4 + (q - 1)) - off
        ty, tq = divmod(target_q_index, 4)
        tq += 1
        months = quarter_months(tq)
        subs = [d for d in team_deals if in_period(d['sub'], ty, months)]
        ita  = [d for d in team_deals if in_period(d['ita'], ty, months)]
        out.append({
            'label'    : f'Q{tq} {str(ty)[-2:]}',
            'year'     : ty, 'quarter': tq,
            'subs'     : len(subs),
            'sub_value': sum(d['amount'] for d in subs),
            'awards'   : len(ita),
            'revenue'  : sum(d['amount'] for d in ita),
        })
    return out

# ─────────────────────────────────────────────
#  Chart helpers — vector via SVG → RLG
# ─────────────────────────────────────────────
CHART_STYLE = {
    'font.family'    : 'DejaVu Sans',
    'axes.spines.top': False,
    'axes.spines.right': False,
    'axes.spines.left': False,
    'axes.facecolor' : 'white',
    'figure.facecolor': 'white',
    'axes.grid'      : False,
    'axes.axisbelow' : True,
    'xtick.color'    : MUTED,
    'ytick.color'    : INK,
    'xtick.labelsize': 8,
    'ytick.labelsize': 8,
    'text.color'     : INK,
}

def fig_to_rlg(fig, target_w_cm, target_h_cm=None):
    buf = io.BytesIO()
    fig.savefig(buf, format='svg', bbox_inches='tight',
                facecolor='white', edgecolor='none')
    plt.close(fig)
    buf.seek(0)
    drawing = svg2rlg(buf)
    if drawing is None:
        return None
    scale_x = (target_w_cm * cm) / drawing.width
    scale_y = (target_h_cm * cm) / drawing.height if target_h_cm else scale_x
    drawing.width  *= scale_x
    drawing.height *= scale_y
    drawing.transform = (scale_x, 0, 0, scale_y, 0, 0)
    return drawing


def chart_hbar(items, w_cm, h_cm, value_fmt=str, max_items=10,
               accent=ACCENT):
    """Horizontal bar chart: list of (label, count) tuples. Plain, no axis.

    Bars are sized for comfortable readability — taller height per bar,
    larger labels, more right-side room for value annotations.
    """
    items = items[:max_items]
    if not items:
        return None
    labels = [k if len(k) <= 32 else k[:30] + '…' for k, _ in items]
    counts = [v for _, v in items]
    n      = len(items)

    with plt.rc_context(CHART_STYLE):
        fig, ax = plt.subplots(figsize=(w_cm/2.54, h_cm/2.54))
        bars = ax.barh(range(n), counts[::-1], color=accent,
                       height=0.65, edgecolor='none')
        ax.set_yticks(range(n))
        ax.set_yticklabels(labels[::-1], fontsize=9, color=INK)
        ax.tick_params(axis='y', length=0, pad=6)
        ax.xaxis.set_visible(False)
        peak = max(counts) if counts else 1
        for bar, cnt in zip(bars, counts[::-1]):
            ax.text(bar.get_width() + peak * 0.02,
                    bar.get_y() + bar.get_height() / 2,
                    value_fmt(cnt), va='center', ha='left',
                    fontsize=9, fontweight='bold', color=NAVY)
        # Extra right padding so the value labels never get clipped
        ax.set_xlim(0, peak * 1.25)
        ax.set_ylim(-0.6, n - 0.4)
        for sp in ax.spines.values():
            sp.set_visible(False)
        fig.tight_layout(pad=0.4)
    return fig_to_rlg(fig, w_cm, h_cm)


def chart_trend(months, w_cm, h_cm):
    """Combo chart: bars = submissions count, line = revenue ($K) — dual axis.

    Designed to avoid label overlap: bar value labels sit just above each bar;
    revenue line labels sit ABOVE the line marker with extra vertical offset
    and only on the most recent + the peak point (rather than every point) so
    they don't clutter when the line and bar tops are close.
    """
    if not months:
        return None
    labels  = [m['label']    for m in months]
    subs    = [m['subs']     for m in months]
    revenue = [m['revenue']  for m in months]
    rev_k   = [r/1000        for r in revenue]
    x       = list(range(len(labels)))

    def _fmt_money_k(v):
        """v is in $K; show as $X.XM if >= 1000, else $XK (rounded)."""
        if v >= 1000: return f'${v/1000:.1f}M'
        if v >= 100:  return f'${v:.0f}K'
        return f'${v:.0f}K'

    with plt.rc_context(CHART_STYLE):
        fig, ax1 = plt.subplots(figsize=(w_cm/2.54, h_cm/2.54))
        # Bars — submissions (soft)
        ax1.bar(x, subs, width=0.55, color=ACCENT, alpha=0.85,
                edgecolor='none', label='Submissions')
        ax1.set_xticks(x)
        ax1.set_xticklabels(labels, fontsize=9, color=INK)
        ax1.tick_params(axis='x', length=0, pad=4)
        ax1.tick_params(axis='y', length=0, pad=2, colors=MUTED, labelsize=8)
        ax1.spines['bottom'].set_color(RULE)
        peak_s = max(subs + [1])
        # Generous headroom (45%) so bar labels don't crash into legend
        ax1.set_ylim(0, peak_s * 1.45)
        # Submission value labels above bars
        for xi, s in zip(x, subs):
            if s > 0:
                ax1.text(xi, s + peak_s * 0.035, str(s), ha='center',
                         fontsize=9, color=NAVY, fontweight='bold')

        # Line — revenue (in $K), navy
        ax2 = ax1.twinx()
        ax2.plot(x, rev_k, color=NAVY, linewidth=1.8,
                 marker='o', markersize=5.5, markerfacecolor=NAVY,
                 markeredgecolor='white', markeredgewidth=1.0,
                 label='Revenue ($, ITA-dated)')
        ax2.tick_params(axis='y', length=0, pad=2, colors=MUTED, labelsize=8)
        for sp in ax2.spines.values():
            sp.set_visible(False)
        peak_r = max(rev_k + [1])
        # Match left-axis headroom for visual balance
        ax2.set_ylim(0, peak_r * 1.45)

        # Annotate revenue points: only the LATEST and the PEAK (if different)
        # Position the label strictly above the marker with comfortable offset
        # so it doesn't crash into the bar-value label below.
        peak_idx = max(range(len(rev_k)), key=lambda i: rev_k[i]) if rev_k else None
        last_idx = len(rev_k) - 1 if rev_k else None
        annotate_idx = set(i for i in (peak_idx, last_idx) if i is not None and rev_k[i] > 0)
        for i in annotate_idx:
            # Vertical offset large enough to clear the bar label
            y_off = peak_r * 0.10
            ax2.annotate(_fmt_money_k(rev_k[i]),
                         xy=(x[i], rev_k[i]),
                         xytext=(x[i], rev_k[i] + y_off),
                         ha='center', fontsize=9, fontweight='bold', color=NAVY)

        # Combined legend, top-left, with extra space above plot
        lines, labs = [], []
        for axx in (ax1, ax2):
            l, lb = axx.get_legend_handles_labels()
            lines += l; labs += lb
        ax1.legend(lines, labs, loc='upper left', fontsize=8,
                   frameon=False, handlelength=1.6, ncol=2,
                   bbox_to_anchor=(0, 1.13))

        fig.tight_layout(pad=0.5)
    return fig_to_rlg(fig, w_cm, h_cm)


# ─────────────────────────────────────────────
#  Paragraph styles
# ─────────────────────────────────────────────
def ps(name, **kw):
    return ParagraphStyle(name, fontName=kw.pop('font', 'Helvetica'), **kw)

ST = {
    'h_title' : ps('ht', font='Helvetica-Bold', fontSize=15, textColor=rl(NAVY), leading=18),
    'h_sub'   : ps('hs', fontSize=8.5, textColor=rl(MUTED), leading=11),
    'sec'     : ps('sl', font='Helvetica-Bold', fontSize=7.5, textColor=rl(NAVY),
                   leading=10, spaceAfter=0),

    'kpi_v'   : ps('kv', font='Helvetica-Bold', fontSize=22, textColor=rl(NAVY),
                   leading=26, alignment=TA_LEFT),
    'kpi_l'   : ps('kl', font='Helvetica-Bold', fontSize=8.5, textColor=rl(MUTED),
                   leading=11, alignment=TA_LEFT),
    'kpi_s'   : ps('ks', fontSize=8, textColor=rl(FAINT), leading=10, alignment=TA_LEFT),

    'th'      : ps('th', font='Helvetica-Bold', fontSize=7, textColor=rl(MUTED), leading=9),
    'th_r'    : ps('thr',font='Helvetica-Bold', fontSize=7, textColor=rl(MUTED),
                   leading=9, alignment=TA_RIGHT),
    'th_c'    : ps('thc',font='Helvetica-Bold', fontSize=7, textColor=rl(MUTED),
                   leading=9, alignment=TA_CENTER),
    'td'      : ps('td', fontSize=8, textColor=rl(INK), leading=11),
    'td_r'    : ps('tdr',fontSize=8, textColor=rl(INK), leading=11, alignment=TA_RIGHT),
    'td_c'    : ps('tdc',fontSize=8, textColor=rl(INK), leading=11, alignment=TA_CENTER),
    'td_v'    : ps('tv', font='Helvetica-Bold', fontSize=8, textColor=rl(NAVY),
                   leading=11, alignment=TA_RIGHT),
    'td_link' : ps('tlk',font='Helvetica-Bold', fontSize=7.5, textColor=rl(ACCENT),
                   leading=10, alignment=TA_CENTER),
    'foot'    : ps('ft', fontSize=7, textColor=rl(FAINT), leading=10, alignment=TA_CENTER),
    'foot_left': ps('fl', fontSize=6.5, textColor=rl(FAINT), leading=9, alignment=TA_LEFT),
    'none'    : ps('no', fontSize=8, textColor=rl(FAINT), leading=11),
}


# ─────────────────────────────────────────────
#  Layout building blocks
# ─────────────────────────────────────────────
def _no_pad():
    return [('TOPPADDING',(0,0),(-1,-1),0), ('BOTTOMPADDING',(0,0),(-1,-1),0),
            ('LEFTPADDING',(0,0),(-1,-1),0), ('RIGHTPADDING',(0,0),(-1,-1),0)]


def header_block(team_name, period_label, UW, subtitle=None, kind='Monthly'):
    # `kind` is passed explicitly by build() — 'Monthly' or 'Quarterly'.
    left_rows = [[Paragraph(f'{team_name}  ·  {kind} Performance Report', ST['h_title'])],
                 [Paragraph(subtitle or
                            f'Public-sector RFP submissions, results, and pipeline  ·  {period_label}',
                            ST['h_sub'])]]
    left = Table(left_rows, colWidths=[UW * 0.70])
    left.setStyle(TableStyle(_no_pad()))

    right_cell = ''
    if os.path.exists(LOGO_PATH):
        lh = 0.85 * cm
        lw = lh * (1205 / 315)
        right_cell = Image(LOGO_PATH, width=lw, height=lh)
    right = Table([[right_cell]], colWidths=[UW * 0.28])
    right.setStyle(TableStyle(_no_pad() + [('ALIGN',(0,0),(-1,-1),'RIGHT'),
                                            ('VALIGN',(0,0),(-1,-1),'MIDDLE')]))

    outer = Table([[left, right]], colWidths=[UW * 0.70, UW * 0.30])
    outer.setStyle(TableStyle(_no_pad() + [('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    return [outer, Spacer(1, 4),
            HRFlowable(width=UW, thickness=1.0, color=rl(NAVY)),
            Spacer(1, 12)]


def sec_label(text, UW):
    return [
        Paragraph(text.upper(), ST['sec']),
        Spacer(1, 2),
        HRFlowable(width=UW, thickness=0.5, color=rl(ACCENT)),
        Spacer(1, 6),
    ]


def sec_label_inline(text, width):
    return [
        Paragraph(text.upper(), ST['sec']),
        Spacer(1, 2),
        HRFlowable(width=width, thickness=0.5, color=rl(ACCENT)),
        Spacer(1, 6),
    ]


def kpi_strip(metrics, UW):
    """4 KPI boxes, each with: big number, label, sub."""
    n = len(metrics)
    cw = [UW / n] * n
    cells = []
    for v, l, s in metrics:
        cell = Table([
            [Paragraph(v, ST['kpi_v'])],
            [Paragraph(l, ST['kpi_l'])],
            [Paragraph(s or '', ST['kpi_s'])],
        ], colWidths=[cw[0] - 8])
        cell.setStyle(TableStyle([
            ('LEFTPADDING', (0,0),(-1,-1), 8),
            ('RIGHTPADDING',(0,0),(-1,-1), 4),
            ('TOPPADDING',  (0,0),(-1,-1), 1),
            ('BOTTOMPADDING',(0,0),(-1,-1),1),
        ]))
        cells.append(cell)
    t = Table([cells], colWidths=cw)
    t.setStyle(TableStyle(_no_pad() + [
        ('LINEABOVE',   (0,0),(-1,-1), 0.6, rl(RULE)),
        ('LINEBELOW',   (0,0),(-1,-1), 0.6, rl(RULE)),
        ('LINEBEFORE',  (1,0),(-1,-1), 0.6, rl(RULE)),
        ('VALIGN',      (0,0),(-1,-1), 'TOP'),
        ('TOPPADDING',  (0,0),(-1,-1), 9),
        ('BOTTOMPADDING',(0,0),(-1,-1),9),
    ]))
    return [t, Spacer(1, 14)]


def two_column(left_widget, right_widget, UW, gap=0.5*cm,
               left_w_ratio=0.5):
    lw = (UW - gap) * left_w_ratio
    rw = (UW - gap) * (1 - left_w_ratio)
    t = Table([[left_widget, '', right_widget]],
              colWidths=[lw, gap, rw])
    t.setStyle(TableStyle(_no_pad() + [('VALIGN',(0,0),(-1,-1),'TOP')]))
    return t


def make_table(header, rows, col_widths, val_cols=None, center_cols=None,
               max_rows=12):
    val_cols    = val_cols    or []
    center_cols = center_cols or []
    rows = rows[:max_rows]
    all_rows = [header] + rows
    styled = []
    for ri, row in enumerate(all_rows):
        sr = []
        for ci, cell in enumerate(row):
            # Pre-built Paragraphs pass through unchanged — keeps custom markup
            if isinstance(cell, Paragraph):
                sr.append(cell)
                continue
            if ri == 0:
                if ci in val_cols:
                    sr.append(Paragraph(str(cell), ST['th_r']))
                elif ci in center_cols:
                    sr.append(Paragraph(str(cell), ST['th_c']))
                else:
                    sr.append(Paragraph(str(cell), ST['th']))
            elif ci in val_cols:
                sr.append(Paragraph(str(cell), ST['td_v']))
            elif ci in center_cols:
                sr.append(Paragraph(str(cell), ST['td_c']))
            else:
                sr.append(Paragraph(str(cell), ST['td']))
        styled.append(sr)
    t = Table(styled, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('LINEBELOW',     (0,0),(-1,0), 0.8, rl(NAVY)),
        ('LINEBELOW',     (0,1),(-1,-1), 0.3, rl(RULE)),
        ('ROWBACKGROUNDS',(0,1),(-1,-1), [rl(WHITE), rl(STRIPE)]),
        ('TOPPADDING',    (0,0),(-1,-1), 5),
        ('BOTTOMPADDING', (0,0),(-1,-1), 5),
        ('LEFTPADDING',   (0,0),(-1,-1), 6),
        ('RIGHTPADDING',  (0,0),(-1,-1), 6),
        ('VALIGN',        (0,0),(-1,-1), 'MIDDLE'),
    ]))
    return t


def deals_block(title, deals, date_field, value_label, UW):
    """Results table (Interview / BAFO / Awards). Deal Name is a clickable
    hyperlink to the deal in HubSpot — no separate link column needed.

    Columns:  Deal Name (RFP) [clickable] | Agency | State | Date | Amount
    """
    if not deals:
        return []
    deals = sorted(deals, key=lambda x: -(x['amount'] or 0))
    rows = []
    for d in deals:
        # Truncate long deal/agency names to avoid pathological wrapping
        raw_name = (d['name'] or '').strip() or '—'
        if len(raw_name) > 75:
            raw_name = raw_name[:73] + '…'
        rfp = (d.get('rfp_no') or '').strip()
        # Show RFP number as-is, but truncate if very long.
        if len(rfp) > 28:
            rfp = rfp[:26] + '…'
        # Decide whether the value already mentions "RFP" / "rfp"; if so don't
        # duplicate the prefix.
        rfp_label = ''
        if rfp:
            rfp_label = (rfp if rfp.lower().startswith(('rfp', 'rfq', 'rf '))
                         else f'RFP {rfp}')

        url = (d.get('link') or '').strip()
        # Build the deal-name cell. Hyperlink the name itself when we have a URL.
        if url:
            name_html = (f'<link href="{xml_escape(url)}" '
                         f'color="{NAVY}">{xml_escape(raw_name)}</link>')
        else:
            name_html = xml_escape(raw_name)
        if rfp_label:
            deal_cell = Paragraph(
                f'{name_html}<br/>'
                f'<font color="{MUTED}" size="7">{xml_escape(rfp_label)}</font>',
                ST['td']
            )
        else:
            deal_cell = Paragraph(name_html, ST['td'])

        agency_raw = (d['agency'] or '').strip() or '—'
        if len(agency_raw) > 55:
            agency_raw = agency_raw[:53] + '…'

        rows.append([
            deal_cell,
            Paragraph(xml_escape(agency_raw), ST['td']),
            Paragraph(xml_escape(d['state'] or '—'), ST['td']),
            Paragraph(str(d.get(date_field)) if d.get(date_field) else '—',
                      ST['td_c']),
            Paragraph(money(d['amount']) if d['amount'] else '—', ST['td_v']),
        ])
    # Column widths — total = UW. Date wide enough to fit "YYYY-MM-DD" inline.
    col_widths = [
        UW * 0.42,   # Deal Name + RFP
        UW * 0.27,   # Agency
        UW * 0.11,   # State
        UW * 0.12,   # Date — must hold "2026-04-30" inline, ~62pt
        UW * 0.08,   # Amount
    ]
    block = []
    block += sec_label(title, UW)
    block.append(make_table(
        ['Deal Name (RFP)', 'Agency', 'State',
         date_field_label(date_field), value_label],
        rows,
        col_widths,
        val_cols=[4],            # Amount column right-aligned
        center_cols=[3],         # Date centered
        max_rows=10,
    ))
    block.append(Spacer(1, 10))
    return block


def date_field_label(field):
    return {
        'iv_date': 'Interview Date',
        'bafo'   : 'BAFO Date',
        'ita'    : 'ITA Date',
        'awarded': 'Awarded Date',
    }.get(field, 'Date')


# ─────────────────────────────────────────────
#  Report builder — single dense dashboard layout
# ─────────────────────────────────────────────
def build(out_path, team_name, team_deals, period, period_label):
    """
    period = {'kind':'month','year':2026,'month':4}
           | {'kind':'quarter','year':2026,'quarter':1,'months':[1,2,3]}
    """
    M  = 1.4 * cm
    UW = A4[0] - 2 * M

    doc = SimpleDocTemplate(
        out_path, pagesize=A4,
        leftMargin=M, rightMargin=M,
        topMargin=M, bottomMargin=1.0 * cm,
    )
    story = []

    # ── Period filter helper ─────────────────────────────────────
    if period['kind'] == 'month':
        months_in_period = [period['month']]
        scope_year       = period['year']
        in_scope = lambda d: in_period(d, scope_year, months_in_period)
        # 6-month trend ending in this month
        trend = six_months(team_deals, scope_year, period['month'])
        trend_subtitle  = '6-Month Trend  ·  Submissions & Revenue'
        period_subtitle = f'Public-sector RFP submissions, results, and pipeline  ·  {period_label}'
        kind_label      = 'Monthly'
    else:  # quarter
        months_in_period = period['months']
        scope_year       = period['year']
        in_scope = lambda d: in_period(d, scope_year, months_in_period)
        trend = four_quarters(team_deals, scope_year, period['quarter'])
        trend_subtitle  = '4-Quarter Trend  ·  Submissions & Revenue'
        period_subtitle = f'Public-sector RFP submissions, results, and pipeline  ·  {period_label}'
        kind_label      = 'Quarterly'

    # ── Period metrics ───────────────────────────────────────────
    sub_deals    = [d for d in team_deals if in_scope(d['sub'])]
    iv_deals     = [d for d in team_deals if in_scope(d['iv_date'])]
    bafo_deals   = [d for d in team_deals if in_scope(d['bafo'])]
    ita_deals    = [d for d in team_deals if in_scope(d['ita'])]   # Awards = ITA date

    cats   = by_category(sub_deals)
    states = by_state(sub_deals, top=8)

    sub_count   = len(sub_deals)
    sub_value   = sum(d['amount'] for d in sub_deals)
    iv_count    = len(iv_deals)
    bafo_count  = len(bafo_deals)
    award_count = len(ita_deals)
    revenue     = sum(d['amount'] for d in ita_deals)

    prev_count = trend[-2]['subs'] if len(trend) >= 2 else 0
    delta      = sub_count - prev_count
    sign       = '+' if delta >= 0 else ''
    prev_label = trend[-2]['label'] if len(trend) >= 2 else 'prior'

    # ── Header ───────────────────────────────────────────────────
    story += header_block(team_name, period_label, UW,
                          subtitle=period_subtitle, kind=kind_label)

    # ── KPI strip — 4 boxes (same shape for monthly + quarterly) ──
    #   Submitted · Interview · Revenue Generated · Awards
    third_kpi = (
        money(revenue) if revenue else '—',
        'Revenue Generated',
        'Sum of $ on ITA-dated deals',
    )
    award_sub = ('Intent to Award this month'
                 if period['kind'] == 'month'
                 else 'Intent to Award this quarter')

    story += kpi_strip([
        (str(sub_count),
         'Submitted',
         f'{money(sub_value) if sub_value else "—"} bid value' +
         (f'  ·  {sign}{delta} vs {prev_label}'
          if len(trend) >= 2 else '')),
        (str(iv_count),
         'Interview',
         'Deals at interview stage'),
        third_kpi,
        (str(award_count),
         'Awards',
         award_sub),
    ], UW)

    # ── Service Category Breakdown — single compact table, no duplicate chart
    if cats:
        story += sec_label('Service Category Breakdown — Bids Submitted', UW)
        rows = []
        total = sum(v['count'] for v in cats.values())
        for k, v in list(cats.items())[:8]:
            pct = v['count']/total*100 if total else 0
            rows.append([k, str(v['count']), f'{pct:.0f}%',
                         money(v['amount']) if v['amount'] else '—'])
        rest = UW - 1.8*cm - 1.6*cm - 2.6*cm
        story.append(make_table(
            ['Category', 'Bids', 'Share', 'Bid Value'],
            rows,
            [rest, 1.8*cm, 1.6*cm, 2.6*cm],
            val_cols=[3], max_rows=8,
        ))
        story.append(Spacer(1, 12))

    # ── Trend (left) + Top States (right) ────────────────────────
    # Trend chart needs more width to spread bars and labels comfortably.
    trend_w  = (UW - 0.5*cm) * 0.60
    states_w = (UW - 0.5*cm) * 0.40
    trend_chart = chart_trend(trend, w_cm=trend_w/cm, h_cm=6.2)
    states_chart = chart_hbar(list(states.items()) if states else [],
                              w_cm=states_w/cm,
                              h_cm=max(5.0, len(states)*0.55 + 1.2),
                              value_fmt=lambda c: str(c),
                              max_items=8) if states else None

    states_label = ('Top States — This Quarter'
                    if period['kind'] == 'quarter' else 'Top States — This Month')
    trend_block = [
        *sec_label_inline(trend_subtitle, trend_w),
        trend_chart if trend_chart else Paragraph('No history.', ST['none']),
    ]
    states_block = [
        *sec_label_inline(states_label, states_w),
        states_chart if states_chart
        else Paragraph('No geographic data.', ST['none']),
    ]
    story.append(two_column(
        Table([[x] for x in trend_block], colWidths=[trend_w]),
        Table([[x] for x in states_block], colWidths=[states_w]),
        UW, gap=0.5*cm, left_w_ratio=0.58,
    ))
    story.append(Spacer(1, 12))

    # ── Detail tables: Interview · BAFO · Awards ────────────────
    # Each block hides itself when there are no deals to show.
    period_word = 'This Quarter' if period['kind'] == 'quarter' else 'This Month'
    story += deals_block(f'Interview Activity — {period_word}',
                         iv_deals,  'iv_date', 'Amount', UW)
    story += deals_block(f'BAFO Activity — {period_word}',
                         bafo_deals, 'bafo',   'Amount', UW)
    story += deals_block(f'Awards (Intent to Award) — {period_word}',
                         ita_deals, 'ita',    'Revenue', UW)

    # Single hint about clickable deal names — shown only if at least one
    # detail table exists.
    if iv_deals or bafo_deals or ita_deals:
        story.append(Paragraph(
            'Deal names above are clickable and open the deal in HubSpot.',
            ST['foot_left']))
        story.append(Spacer(1, 6))

    # ── Footer ──────────────────────────────────────────────────
    story.append(HRFlowable(width=UW, thickness=0.4, color=rl(RULE)))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f'CAPS  ·  {team_name}  ·  {period_label}  ·  '
        f'Generated {date.today().strftime("%d %B %Y")}  ·  Internal & Confidential',
        ST['foot']))

    doc.build(story)
    kb = os.path.getsize(out_path) // 1024
    print(f'  ✓  {os.path.basename(out_path)}  ({kb} KB)')


# ─────────────────────────────────────────────
#  Entry point
# ─────────────────────────────────────────────
def write_manifest():
    """Scan SCRIPT_DIR for all generated report PDFs and emit manifest.json
    so the dashboard UI can populate dropdowns dynamically."""
    monthly = {}   # year -> {period_folder: {label, teams: [...]}}
    quarterly = {} # year -> {period_folder: {label, teams: [...]}}

    for entry in os.listdir(SCRIPT_DIR):
        full = os.path.join(SCRIPT_DIR, entry)
        if not os.path.isdir(full):
            continue
        # Quarterly folder: Q1-2026
        q_match = re.match(r'^Q([1-4])-(\d{4})$', entry)
        m_match = re.match(r'^([A-Z][a-z]+)-(\d{4})$', entry)
        teams_in_folder = []
        for fn in sorted(os.listdir(full)):
            if not fn.endswith('.pdf'):
                continue
            teams_in_folder.append(fn)
        if not teams_in_folder:
            continue
        if q_match:
            q = int(q_match.group(1)); yr = int(q_match.group(2))
            quarterly.setdefault(str(yr), {})[entry] = {
                'label': f'Q{q} {yr}',
                'sort':  yr * 10 + q,
                'files': teams_in_folder,
            }
        elif m_match:
            month_name = m_match.group(1); yr = int(m_match.group(2))
            try:
                mo = datetime.strptime(month_name, '%B').month
            except ValueError:
                continue
            monthly.setdefault(str(yr), {})[entry] = {
                'label': f'{month_name} {yr}',
                'sort':  yr * 100 + mo,
                'files': teams_in_folder,
            }

    manifest = {
        'generated_at_utc': datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%SZ'),
        'monthly':   monthly,
        'quarterly': quarterly,
        'teams':     list(TEAMS.keys()) + ['Company Summary'],
    }
    out_json = os.path.join(SCRIPT_DIR, 'manifest.json')
    with open(out_json, 'w') as f:
        json.dump(manifest, f, indent=2)
    # Also emit a .js file that sets window.CAPS_REPORTS_MANIFEST.
    # This lets the dashboard load the manifest via a <script> tag instead of
    # fetch() — works under file:// (no CORS) AND on GitHub Pages.
    out_js = os.path.join(SCRIPT_DIR, 'manifest.js')
    with open(out_js, 'w') as f:
        f.write('/* Auto-generated by generate_monthly_reports.py — do not edit manually */\n')
        f.write('window.CAPS_REPORTS_MANIFEST = ' + json.dumps(manifest, indent=2) + ';\n')
    n_m = sum(len(v) for v in monthly.values())
    n_q = sum(len(v) for v in quarterly.values())
    print(f'  · manifest.json + manifest.js updated  ({n_m} monthly, {n_q} quarterly periods)')


def generate_for_period(period, period_label, folder, tag, all_deals):
    """Generate all team PDFs + Company Summary for the given period."""
    out_dir = os.path.join(SCRIPT_DIR, folder)
    os.makedirs(out_dir, exist_ok=True)
    print(f'  Output: {out_dir}\n')
    for team_name in TEAMS:
        team_deals = [d for d in all_deals if d['team'] == team_name]
        build(os.path.join(out_dir, f'{team_name.replace(" ","_")}_{tag}.pdf'),
              team_name, team_deals, period, period_label)
    build(os.path.join(out_dir, f'Company_Summary_{tag}.pdf'),
          'Company Summary', all_deals, period, period_label)
    print(f'\n  Done — {out_dir}')
    for f in sorted(os.listdir(out_dir)):
        if f.endswith('.pdf'):
            print(f'  · {f}  ({os.path.getsize(os.path.join(out_dir,f))//1024} KB)')


def main():
    import re as _re_check  # ensure re is imported (used in write_manifest)
    ap = argparse.ArgumentParser()
    ap.add_argument('--month',   help='YYYY-MM (e.g. 2026-04)')
    ap.add_argument('--quarter', help='YYYY-Q[1-4] (e.g. 2026-Q1)')
    args = ap.parse_args()

    if not os.path.exists(EXCEL_FILE):
        print(f'ERROR: Excel not found: {EXCEL_FILE}'); sys.exit(1)
    all_deals = load_data()

    if args.quarter:
        m = re.match(r'^(\d{4})-Q([1-4])$', args.quarter)
        if not m:
            print('ERROR: --quarter must be YYYY-Q[1-4]'); sys.exit(1)
        yr, q = int(m.group(1)), int(m.group(2))
        period = {'kind':'quarter','year':yr,'quarter':q,'months':quarter_months(q)}
        period_label = f'Q{q} {yr}'
        folder = f'Q{q}-{yr}'
        tag    = f'Q{q}_{yr}'
        print(f'\n  CAPS Quarterly Reports — {period_label}')
        generate_for_period(period, period_label, folder, tag, all_deals)
    elif args.month:
        t = datetime.strptime(args.month, '%Y-%m').date()
        yr, mo = t.year, t.month
        period = {'kind':'month','year':yr,'month':mo}
        period_label = date(yr, mo, 1).strftime('%B %Y')
        folder = date(yr, mo, 1).strftime('%B-%Y')
        tag    = date(yr, mo, 1).strftime('%b_%Y')
        print(f'\n  CAPS Monthly Reports — {period_label}')
        generate_for_period(period, period_label, folder, tag, all_deals)
    else:
        # Default: last completed month
        p = date.today() - relativedelta(months=1)
        yr, mo = p.year, p.month
        period = {'kind':'month','year':yr,'month':mo}
        period_label = date(yr, mo, 1).strftime('%B %Y')
        folder = date(yr, mo, 1).strftime('%B-%Y')
        tag    = date(yr, mo, 1).strftime('%b_%Y')
        print(f'\n  CAPS Monthly Reports — {period_label}')
        generate_for_period(period, period_label, folder, tag, all_deals)

    # Always refresh the manifest after generating any report
    print()
    write_manifest()


if __name__ == '__main__':
    main()
