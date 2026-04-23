"""
Run this after updating CAPS_RFP_Dashboard_Dataset.xlsx
It regenerates js/data.js so the dashboard picks up the latest data.

Usage:  python refresh-data.py
"""
import json, os, sys
from datetime import datetime, date

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...")
    os.system(sys.executable + " -m pip install openpyxl -q")
    import openpyxl

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PARENT_DIR = os.path.dirname(SCRIPT_DIR)
EXCEL_NAME = "CAPS_RFP_Dashboard_Dataset.xlsx"
# Look for Excel in parent folder first (where you edit it), then in dashboard/data/
EXCEL_PATH_PRIMARY = os.path.join(PARENT_DIR, EXCEL_NAME)
EXCEL_PATH_FALLBACK = os.path.join(SCRIPT_DIR, "data", EXCEL_NAME)
EXCEL_PATH = EXCEL_PATH_PRIMARY if os.path.exists(EXCEL_PATH_PRIMARY) else EXCEL_PATH_FALLBACK
COPY_DEST = os.path.join(SCRIPT_DIR, "data", EXCEL_NAME)
OUTPUT_PATH = os.path.join(SCRIPT_DIR, "js", "data.js")

def serialize(val):
    if val is None or val == "":
        return ""
    if isinstance(val, (datetime, date)):
        return val.isoformat()
    if isinstance(val, (int, float)):
        return val
    return str(val)

def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: Excel file not found at:\n  {EXCEL_PATH}")
        sys.exit(1)

    print(f"Reading: {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH)

    # Save the workbook back so the README timestamp is persisted in the Excel file
    wb.save(EXCEL_PATH)

    # Copy to dashboard/data/ so it stays in sync for GitHub
    import shutil
    if EXCEL_PATH != COPY_DEST:
        shutil.copy2(EXCEL_PATH, COPY_DEST)
        print(f"Copied to: {COPY_DEST}")

    # Stamp current run time as last updated (always fresh, never stale)
    last_updated = datetime.now().strftime("%B %d, %Y %I:%M %p")

    # Write the new timestamp back into the README sheet so Excel stays in sync
    if "README" in wb.sheetnames:
        ws_readme = wb["README"]
        for row in ws_readme.iter_rows():
            if row[0].value and isinstance(row[0].value, str) and row[0].value.startswith("Last Updated:"):
                row[0].value = f"Last Updated: {last_updated}"
                break

    def parse_sheet(sheet_name):
        sheet = wb[sheet_name]
        headers = [cell.value for cell in sheet[1]]
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            record = {}
            for h, v in zip(headers, row):
                if h:
                    record[h.strip()] = serialize(v)
            rows.append(record)
        return rows

    # Parse RFP Data
    records = parse_sheet("RFP Data")

    # Parse Awards sheet (all-time awards, no date filter)
    awards_records = parse_sheet("Awards") if "Awards" in wb.sheetnames else []

    data = {"lastUpdated": last_updated, "records": records}
    js = "/* Auto-generated from Excel — do not edit manually */\n"
    js += f"/* Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')} */\n"
    js += "window.CAPS_EMBEDDED_DATA = " + json.dumps(data, separators=(",", ":")) + ";\n"
    js += "window.CAPS_AWARDS_DATA = " + json.dumps(awards_records, separators=(",", ":")) + ";\n"

    with open(OUTPUT_PATH, "w") as f:
        f.write(js)

    print(f"Done! {len(records)} records written to js/data.js")
    print(f"Awards sheet: {len(awards_records)} records written to CAPS_AWARDS_DATA")
    print(f"Last updated: {last_updated}")

    # ─── Target-states page: per-state summary (Oct 2025+) ───
    # The "Top Target States" dashboard filters RFP Data to the target states
    # with bidClosingDate / submissionDate on or after Oct 1 2025. Print a summary
    # so you can verify every target state has data after each refresh — and catch
    # typos in the Agency State column.
    TARGET_STATES = ["California", "New York", "Texas", "Florida"]
    OCT_2025 = datetime(2025, 10, 1)

    def _to_dt(v):
        if not v:
            return None
        if isinstance(v, datetime):
            return v
        if isinstance(v, date):
            return datetime(v.year, v.month, v.day)
        try:
            # ISO string from serialize()
            return datetime.fromisoformat(str(v).replace("Z", ""))
        except (ValueError, TypeError):
            return None

    state_counts = {s: {"bids": 0, "won": 0} for s in TARGET_STATES}
    for r in records:
        state = (r.get("Agency State") or "").strip()
        if state not in state_counts:
            continue
        ref_dt = _to_dt(r.get("Bid Closing Date")) or _to_dt(r.get("Submission Date"))
        if not ref_dt or ref_dt < OCT_2025:
            continue
        state_counts[state]["bids"] += 1
        if (r.get("Stage") or "").strip() == "Closed Won":
            state_counts[state]["won"] += 1

    print("\nTop Target States page — bids since Oct 2025:")
    for s in TARGET_STATES:
        c = state_counts[s]
        flag = "" if c["bids"] > 0 else "  ⚠ no data"
        print(f"  {s:<12}  {c['bids']:>4} bids  ({c['won']} won){flag}")

if __name__ == "__main__":
    main()
