#!/usr/bin/env python3
"""
validate_and_refresh.py — 10-day deep-check & validation script
================================================================

Two modes:

1. **Validate mode** (no fresh data needed):
       python3 validate_and_refresh.py --validate

   Sanity-checks the EXISTING snapshot files and Excel:
   - Row counts per sheet
   - Duplicate hs_object_id detection
   - Missing state count
   - Date format consistency
   - Stage value coverage
   - Orphan deals (in Excel but not in JSON, or vice versa)

2. **Diff & refresh mode** (full field-level reconciliation):
       python3 validate_and_refresh.py --diff \
           --fresh-rfp /path/to/rfp_fresh.json \
           --fresh-awards /path/to/awards_fresh.json \
           [--apply]

   Compares fresh HubSpot data against the snapshot field-by-field for ALL
   fields and prints a detailed change report (added / removed / modified
   deals, with old → new values per field). With --apply, overwrites the
   snapshot files. Run update_excel_v2.py + refresh-data.py afterward to
   propagate changes to the dashboard.

How to get fresh-rfp.json / fresh-awards.json:
   This script is plain Python — it cannot call HubSpot MCP itself. Have
   Claude (or you, manually via the HubSpot API) pull all RFP-stage and
   Awards-stage deals fresh and combine into two arrays. Save as JSON in
   the same {id, properties: {...}} shape as rfp_deals_all.json. See
   §17 of DASHBOARD_README.md for the recommended pull procedure.

Recommended cadence:
   - Daily: run the lightweight refresh (NEW + REMOVED only) — see README §17.
   - Every 10 days: run this script with --diff --apply to catch any silent
     amount/date/category edits that the daily refresh misses.

Usage examples:

   # Just sanity-check the data
   python3 validate_and_refresh.py --validate

   # See what would change (dry-run)
   python3 validate_and_refresh.py --diff \
       --fresh-rfp ../outputs/rfp_fresh.json \
       --fresh-awards ../outputs/awards_fresh.json

   # See what would change AND apply it
   python3 validate_and_refresh.py --diff \
       --fresh-rfp ../outputs/rfp_fresh.json \
       --fresh-awards ../outputs/awards_fresh.json \
       --apply

Exit codes:
   0  success (no issues / changes)
   1  validation issues found
   2  changes detected (dry-run mode without --apply)
   3  bad arguments / missing input files
"""

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta

try:
    from zoneinfo import ZoneInfo
    EASTERN = ZoneInfo("America/New_York")
except ImportError:
    EASTERN = timezone(timedelta(hours=-4))  # EDT fallback

import openpyxl


# ─────────────────────────────────────────────
# Paths — discovered relative to this file
# ─────────────────────────────────────────────
SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))           # .../dashboard/
PROJECT_DIR = os.path.dirname(SCRIPT_DIR)                          # .../CAPS UPDATE DASHBOARD/
EXCEL_PATH  = os.path.join(PROJECT_DIR, 'CAPS_RFP_Dashboard_Dataset.xlsx')

# Snapshot JSONs live alongside the per-session pull files. We try a few
# common locations so the script works whether you run it from the project
# root or from dashboard/.
SNAPSHOT_CANDIDATES = [
    # New location — co-located with this script
    os.path.join(SCRIPT_DIR, 'data', 'rfp_deals_all.json'),
    # Older convention — in an outputs/ folder near the project
    os.path.join(PROJECT_DIR, 'outputs', 'rfp_deals_all.json'),
    # Linux session sandbox path (when run from Cowork / Claude Code session)
    '/sessions/blissful-trusting-clarke/mnt/outputs/rfp_deals_all.json',
]


def find_snapshot_dir():
    for candidate in SNAPSHOT_CANDIDATES:
        if os.path.exists(candidate):
            return os.path.dirname(candidate)
    return None


# ─────────────────────────────────────────────
# Stage map (must mirror update_excel_v2.py)
# ─────────────────────────────────────────────
STAGE_MAP = {
    "presentationscheduled": "Submitted",
    "1620129473":             "Interview",
    "2485737153":             "Intent to Award",
    "closedwon":              "Closed Won",
    "closedlost":             "Closed Lost",
    "2203296493":             "Terminated",
    "2766010076":             "RFx Cancelled",
}
RFP_STAGES   = set(STAGE_MAP.keys())
AWARD_STAGES = {"closedwon", "2485737153"}

# Fields we compare in field-level diff. Order matters for nice output.
DIFF_FIELDS = [
    'dealname', 'dealstage', 'amount', 'closedate', 'submission_date',
    'agency', 'rfp_number', 'service_category__cloned_', 'submission_mode',
    'hubspot_owner_id', 'interview_type', 'interview_date_time', 'bafo_date',
    'intent_to_awarded_date', 'tentatively_awarded_date', 'awarded_date',
    'current_status_of_award', 'closed_won_reason', 'reason_of_close_lost',
    'delivery_needed', 'createdate',
]

# ANSI colors (degrade gracefully on Windows / non-tty)
USE_COLOR = sys.stdout.isatty()
def _c(code, s):
    return f"\033[{code}m{s}\033[0m" if USE_COLOR else s
RED    = lambda s: _c('31', s)
GREEN  = lambda s: _c('32', s)
YELLOW = lambda s: _c('33', s)
BLUE   = lambda s: _c('34', s)
CYAN   = lambda s: _c('36', s)
BOLD   = lambda s: _c('1',  s)
DIM    = lambda s: _c('2',  s)


# ─────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────
def now_eastern_str():
    return datetime.now(EASTERN).strftime("%Y-%m-%d %H:%M %Z")


def load_json(path):
    if not os.path.exists(path):
        sys.stderr.write(f"ERROR: file not found: {path}\n")
        sys.exit(3)
    with open(path) as f:
        try:
            data = json.load(f)
        except json.JSONDecodeError as e:
            sys.stderr.write(f"ERROR: invalid JSON in {path}: {e}\n")
            sys.exit(3)
    if isinstance(data, dict) and 'results' in data:
        # Tolerate raw HubSpot API response shape
        data = data['results']
    if not isinstance(data, list):
        sys.stderr.write(f"ERROR: {path} must be a JSON array of deals\n")
        sys.exit(3)
    return data


def by_id(deals):
    out = {}
    for d in deals:
        did = str(d.get('id') or d.get('properties', {}).get('hs_object_id', ''))
        if did:
            out[did] = d
    return out


def truncate(s, n=60):
    s = str(s) if s is not None else ''
    return s if len(s) <= n else s[:n - 1] + '…'


def normalise(v):
    """Treat None / empty string / whitespace as equivalent. Strip strings."""
    if v is None:
        return ''
    return str(v).strip()


# ─────────────────────────────────────────────
# Validation mode
# ─────────────────────────────────────────────
def cmd_validate():
    print(BOLD(f"\n  Dashboard Data Validation  ·  {now_eastern_str()}"))
    print("  " + "─" * 70)

    snapshot_dir = find_snapshot_dir()
    if not snapshot_dir:
        sys.stderr.write("\n  Could not locate snapshot files (rfp_deals_all.json / awards_deals_all.json).\n")
        sys.stderr.write("  Tried:\n")
        for c in SNAPSHOT_CANDIDATES:
            sys.stderr.write(f"    - {c}\n")
        sys.exit(3)
    print(f"  Snapshot dir : {snapshot_dir}")
    print(f"  Excel        : {EXCEL_PATH}")
    print()

    issues = []

    # 1. Snapshot JSON files
    rfp_path    = os.path.join(snapshot_dir, 'rfp_deals_all.json')
    awards_path = os.path.join(snapshot_dir, 'awards_deals_all.json')
    rfp    = load_json(rfp_path)
    awards = load_json(awards_path)

    print(BOLD("  [1] Snapshot counts"))
    print(f"      rfp_deals_all.json    : {len(rfp)} deals")
    print(f"      awards_deals_all.json : {len(awards)} deals")

    # 2. Duplicate IDs
    print()
    print(BOLD("  [2] Duplicate hs_object_id check"))
    for label, src in [('RFP', rfp), ('Awards', awards)]:
        ids = [str(d.get('id') or d.get('properties', {}).get('hs_object_id', ''))
               for d in src]
        dup = [i for i, c in Counter(ids).items() if c > 1]
        if dup:
            issues.append(f"{label}: {len(dup)} duplicate IDs")
            print(f"      {label:<10} {RED(str(len(dup)))} duplicates: {dup[:5]}{'…' if len(dup)>5 else ''}")
        else:
            print(f"      {label:<10} {GREEN('no duplicates')}")

    # 3. Stage value coverage
    print()
    print(BOLD("  [3] Stage distribution (RFP)"))
    stages = Counter()
    unknown_stages = []
    for d in rfp:
        s = d.get('properties', {}).get('dealstage', '')
        stages[s] += 1
        if s not in STAGE_MAP:
            unknown_stages.append(s)
    for s, n in stages.most_common():
        label = STAGE_MAP.get(s, RED(f"UNKNOWN: {s}"))
        print(f"      {n:>4}  {label}")
    if unknown_stages:
        issues.append(f"Found {len(unknown_stages)} deals with unknown stage")

    # 4. Date format consistency
    print()
    print(BOLD("  [4] Date format consistency"))
    DATE_FIELDS = ['submission_date', 'closedate', 'interview_date_time',
                   'bafo_date', 'intent_to_awarded_date', 'awarded_date',
                   'tentatively_awarded_date', 'createdate']
    bad_dates = 0
    for d in rfp + awards:
        p = d.get('properties', {})
        for f in DATE_FIELDS:
            v = p.get(f)
            if not v:
                continue
            s = str(v)
            # Accept either YYYY-MM-DD or full ISO 8601
            if not (re.match(r'^\d{4}-\d{2}-\d{2}$', s)
                    or re.match(r'^\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}', s)):
                bad_dates += 1
    if bad_dates:
        issues.append(f"{bad_dates} date fields have unexpected format")
        print(f"      {RED(str(bad_dates))} unexpected date formats")
    else:
        print(f"      {GREEN('all dates well-formed')}")

    # 5. Missing state count
    print()
    print(BOLD("  [5] Excel — Agency State coverage"))
    if not os.path.exists(EXCEL_PATH):
        issues.append("Excel file missing")
        print(f"      {RED('Excel not found')}")
    else:
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        for sheet_name in ['RFP Data', 'Awards']:
            if sheet_name not in wb.sheetnames:
                issues.append(f"Excel sheet missing: {sheet_name}")
                continue
            ws = wb[sheet_name]
            h = [c.value for c in ws[1]]
            if 'Agency State' not in h:
                continue
            ix_st = h.index('Agency State')
            ix_id = h.index('HubSpot ID') if 'HubSpot ID' in h else None
            ix_ag = h.index('Agency')     if 'Agency'     in h else None
            total = with_state = no_agency_no_state = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row[ix_id]:
                    continue
                total += 1
                state = (row[ix_st] or '').strip() if ix_st is not None else ''
                agency = (row[ix_ag] or '').strip() if ix_ag is not None else ''
                if state:
                    with_state += 1
                elif not agency:
                    no_agency_no_state += 1
            pct = with_state / total * 100 if total else 0
            color = GREEN if pct >= 95 else (YELLOW if pct >= 90 else RED)
            print(f"      {sheet_name:<10} {with_state}/{total} have state  "
                  f"({color(f'{pct:.1f}%')})  ·  {no_agency_no_state} have blank agency in HubSpot")
            if pct < 90:
                issues.append(f"{sheet_name} state coverage below 90%")

    # 6. Cross-check: Excel rows vs JSON snapshot
    print()
    print(BOLD("  [6] Cross-check Excel vs snapshot JSON"))
    if os.path.exists(EXCEL_PATH):
        wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
        for sheet_name, snap in [('RFP Data', rfp), ('Awards', awards)]:
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            h = [c.value for c in ws[1]]
            ix_id = h.index('HubSpot ID') if 'HubSpot ID' in h else None
            if ix_id is None:
                continue
            excel_ids = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if row[ix_id]:
                    excel_ids.add(str(row[ix_id]))
            snap_ids  = set(str(d.get('id') or d.get('properties', {}).get('hs_object_id', '')) for d in snap)
            in_excel_only  = excel_ids - snap_ids
            in_snap_only   = snap_ids - excel_ids
            ok = not in_excel_only and not in_snap_only
            if ok:
                print(f"      {sheet_name:<10} {GREEN('IDs match')}  ({len(snap_ids)} deals)")
            else:
                if in_excel_only:
                    issues.append(f"{sheet_name}: {len(in_excel_only)} IDs in Excel but not in JSON")
                    print(f"      {sheet_name:<10} {RED('mismatch')} — {len(in_excel_only)} in Excel only, {len(in_snap_only)} in JSON only")

    # Summary
    print()
    print("  " + "─" * 70)
    if issues:
        print(BOLD(RED(f"  ✗ Validation found {len(issues)} issue(s):")))
        for issue in issues:
            print(f"      · {issue}")
        sys.exit(1)
    else:
        print(BOLD(GREEN(f"  ✓ All validations passed.")))
        sys.exit(0)


# ─────────────────────────────────────────────
# Diff & refresh mode
# ─────────────────────────────────────────────
def cmd_diff(fresh_rfp_path, fresh_awards_path, apply_changes, report_path):
    print(BOLD(f"\n  10-Day Deep Check  ·  Field-Level Diff  ·  {now_eastern_str()}"))
    print("  " + "─" * 70)

    snapshot_dir = find_snapshot_dir()
    if not snapshot_dir:
        sys.stderr.write("ERROR: snapshot dir not found\n")
        sys.exit(3)

    snap_rfp_path    = os.path.join(snapshot_dir, 'rfp_deals_all.json')
    snap_awards_path = os.path.join(snapshot_dir, 'awards_deals_all.json')

    snap_rfp    = load_json(snap_rfp_path)
    snap_awards = load_json(snap_awards_path)
    fresh_rfp    = load_json(fresh_rfp_path)
    fresh_awards = load_json(fresh_awards_path)

    print(f"  Snapshot:  {len(snap_rfp)} RFP, {len(snap_awards)} Awards")
    print(f"  Fresh:     {len(fresh_rfp)} RFP, {len(fresh_awards)} Awards")

    report_lines = []
    def out(line=''):
        print(line)
        report_lines.append(re.sub(r'\033\[[0-9;]+m', '', line))  # strip ANSI

    total_changes = {'added': 0, 'removed': 0, 'modified': 0,
                     'modifications': 0}

    for label, snap, fresh in [('RFP', snap_rfp, fresh_rfp),
                                ('Awards', snap_awards, fresh_awards)]:
        out()
        out(BOLD(CYAN(f"  ═══ {label} sheet ═══")))
        snap_by  = by_id(snap)
        fresh_by = by_id(fresh)

        added    = sorted(set(fresh_by) - set(snap_by))
        removed  = sorted(set(snap_by)  - set(fresh_by))
        common   = set(snap_by) & set(fresh_by)

        modified = []  # list of (id, deal_name, [(field, old, new), ...])
        for did in sorted(common):
            old_p = snap_by[did].get('properties', {})
            new_p = fresh_by[did].get('properties', {})
            diffs = []
            for f in DIFF_FIELDS:
                if normalise(old_p.get(f)) != normalise(new_p.get(f)):
                    diffs.append((f, old_p.get(f, ''), new_p.get(f, '')))
            if diffs:
                modified.append((did, new_p.get('dealname', ''), diffs))

        out(f"    Added    : {GREEN(str(len(added)))}")
        out(f"    Removed  : {RED(str(len(removed)))}")
        out(f"    Modified : {YELLOW(str(len(modified)))}")

        total_changes['added']    += len(added)
        total_changes['removed']  += len(removed)
        total_changes['modified'] += len(modified)

        if added:
            out()
            out(GREEN("    + ADDED:"))
            for did in added[:50]:
                p = fresh_by[did].get('properties', {})
                out(f"      + {did}  {truncate(p.get('dealname'), 60)}  "
                    f"[{STAGE_MAP.get(p.get('dealstage', ''), p.get('dealstage', ''))}]")
            if len(added) > 50:
                out(f"      … and {len(added) - 50} more")

        if removed:
            out()
            out(RED("    − REMOVED:"))
            for did in removed[:50]:
                p = snap_by[did].get('properties', {})
                out(f"      − {did}  {truncate(p.get('dealname'), 60)}")
            if len(removed) > 50:
                out(f"      … and {len(removed) - 50} more")

        if modified:
            out()
            out(YELLOW("    ~ MODIFIED:"))
            for did, name, diffs in modified[:100]:
                out(f"      ~ {did}  {truncate(name, 60)}")
                for f, ov, nv in diffs:
                    total_changes['modifications'] += 1
                    out(f"          {f:<28}: {DIM(truncate(ov, 35))}  →  {truncate(nv, 35)}")

    # Summary
    out()
    out("  " + "─" * 70)
    out(f"  Total changes : "
        f"{GREEN(str(total_changes['added'])+' added')}, "
        f"{RED(str(total_changes['removed'])+' removed')}, "
        f"{YELLOW(str(total_changes['modified'])+' modified')} "
        f"({total_changes['modifications']} field changes)")

    # Write a markdown report alongside the script for record-keeping
    if report_path is None:
        report_path = os.path.join(SCRIPT_DIR, 'last_diff_report.md')
    with open(report_path, 'w') as f:
        f.write(f"# Field-level diff report\n\n")
        f.write(f"Generated: {now_eastern_str()}\n\n")
        f.write("```\n")
        f.write("\n".join(report_lines))
        f.write("\n```\n")
    out()
    out(f"  Report written: {report_path}")

    # Apply
    no_changes = (total_changes['added'] == 0
                  and total_changes['removed'] == 0
                  and total_changes['modified'] == 0)
    if no_changes:
        out(BOLD(GREEN("  ✓ Snapshot is already current — no changes to apply.")))
        sys.exit(0)

    if apply_changes:
        # Overwrite snapshots with fresh data, sorted to keep diffs readable
        snap_rfp_sorted = sorted(
            fresh_rfp,
            key=lambda d: d.get('properties', {}).get('closedate') or '')
        snap_awards_sorted = sorted(
            fresh_awards,
            key=lambda d: d.get('properties', {}).get('intent_to_awarded_date') or '')
        with open(snap_rfp_path, 'w') as f:
            json.dump(snap_rfp_sorted, f, indent=2)
        with open(snap_awards_path, 'w') as f:
            json.dump(snap_awards_sorted, f, indent=2)
        out(BOLD(GREEN("  ✓ Snapshot files updated.")))
        out(f"     {snap_rfp_path}")
        out(f"     {snap_awards_path}")
        out()
        out(BOLD("  Next steps:"))
        out("     1. Review the changes above (or in last_diff_report.md).")
        out("     2. Run:  python3 update_excel_v2.py")
        out("              cd dashboard && python3 refresh-data.py")
        out("     3. Hard-refresh the dashboard in your browser.")
        sys.exit(0)
    else:
        out(BOLD(YELLOW("  ⚠ Dry-run mode — no files modified.")))
        out("     Re-run with --apply to write the snapshot.")
        sys.exit(2)


# ─────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────
def main():
    p = argparse.ArgumentParser(
        description="Validate dashboard data + 10-day deep refresh against fresh HubSpot pulls.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    g = p.add_mutually_exclusive_group(required=True)
    g.add_argument('--validate', action='store_true',
                   help='Sanity-check existing snapshot files + Excel.')
    g.add_argument('--diff', action='store_true',
                   help='Field-level diff against fresh HubSpot data.')

    p.add_argument('--fresh-rfp', help='Path to fresh RFP JSON (required with --diff).')
    p.add_argument('--fresh-awards', help='Path to fresh Awards JSON (required with --diff).')
    p.add_argument('--apply', action='store_true',
                   help='With --diff, overwrite snapshot files with fresh data.')
    p.add_argument('--report', help='Path for markdown report (default: dashboard/last_diff_report.md).')

    args = p.parse_args()

    if args.validate:
        cmd_validate()
    elif args.diff:
        if not args.fresh_rfp or not args.fresh_awards:
            sys.stderr.write("ERROR: --diff requires --fresh-rfp and --fresh-awards\n")
            sys.exit(3)
        cmd_diff(args.fresh_rfp, args.fresh_awards, args.apply, args.report)


if __name__ == '__main__':
    main()
